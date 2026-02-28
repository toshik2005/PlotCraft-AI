import argparse
import os
import re
from typing import Iterable, Optional

from openpyxl import load_workbook


def _clean(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+\n", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    s = re.sub(r"[ \t]{2,}", " ", s)
    return s.strip()


def _iter_rows(ws) -> Iterable[list]:
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def _pick_text_column(headers: list[str], sample_rows: list[list]) -> int:
    normalized = [str(h or "").strip().lower() for h in headers]
    keyword_order = ["story", "text", "content", "body", "creepypasta", "pasta", "description"]

    def avg_len(i: int) -> float:
        lengths: list[int] = []
        for r in sample_rows:
            if i >= len(r):
                continue
            v = r[i]
            if isinstance(v, str):
                vv = _clean(v)
                if vv:
                    lengths.append(len(vv))
        return (sum(lengths) / len(lengths)) if lengths else 0.0

    # Prefer "text-like" columns, but pick the one that actually contains long text.
    keyword_candidates: list[int] = []
    for kw in keyword_order:
        for i, h in enumerate(normalized):
            if kw in h:
                keyword_candidates.append(i)

    if keyword_candidates:
        return max(keyword_candidates, key=avg_len)

    # Fallback: choose column with largest average string length in sample
    return max(range(len(headers)), key=avg_len)


def main() -> None:
    plotcraft_base = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    repo_root = os.path.abspath(os.path.join(plotcraft_base, "..", ".."))

    parser = argparse.ArgumentParser(description="Convert creepypastas.xlsx into cleaned text corpus.")
    parser.add_argument(
        "--input",
        default=os.path.join(repo_root, "creepypastas.xlsx"),
        help="Path to creepypastas.xlsx",
    )
    parser.add_argument(
        "--output",
        default=os.path.join(plotcraft_base, "data", "processed", "cleaned_fixed_horror.txt"),
        help="Output cleaned text path.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name. Defaults to first sheet.",
    )
    parser.add_argument(
        "--text_column",
        default=None,
        help="Optional column name for story text. If omitted, auto-detected.",
    )
    parser.add_argument("--min_chars", type=int, default=200, help="Drop rows shorter than this.")
    parser.add_argument("--max_rows", type=int, default=0, help="0 = all rows, else cap for quick runs.")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        raise FileNotFoundError(f"Excel file not found: {args.input}")

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    rows_iter = _iter_rows(ws)
    headers = [str(h or "").strip() for h in next(rows_iter)]
    sample_rows: list[list] = []
    for _ in range(200):
        try:
            sample_rows.append(next(rows_iter))
        except StopIteration:
            break

    # Re-create iterator (read_only mode isn't seekable); reload sheet
    wb.close()
    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows_iter = _iter_rows(ws)
    _ = next(rows_iter)  # skip headers

    col_idx: int
    if args.text_column:
        normalized = [h.strip().lower() for h in headers]
        target = args.text_column.strip().lower()
        if target not in normalized:
            raise ValueError(f"Column '{args.text_column}' not found. Available: {headers}")
        col_idx = normalized.index(target)
    else:
        col_idx = _pick_text_column(headers, sample_rows)

    kept = 0
    seen = 0
    with open(args.output, "w", encoding="utf-8") as out:
        for row in rows_iter:
            if args.max_rows and seen >= args.max_rows:
                break
            seen += 1
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if not isinstance(v, str):
                continue
            text = _clean(v)
            if len(text) < args.min_chars:
                continue
            out.write(text)
            out.write("\n\n")
            kept += 1

    wb.close()
    print("Done.")
    print("Sheet:", ws.title)
    print("Detected text column:", headers[col_idx] if col_idx < len(headers) else col_idx)
    print("Rows scanned:", seen)
    print("Stories kept:", kept)
    print("Saved to:", args.output)

    if kept == 0:
        raise RuntimeError(
            "No stories were extracted. Your story column is probably not the one auto-detected. "
            "Re-run with: --text_column body (or the correct column name). "
            f"Available columns: {headers}"
        )


if __name__ == "__main__":
    main()
